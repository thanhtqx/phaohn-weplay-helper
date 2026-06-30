import csv
import io


def looks_like_header(cells: list[str]) -> bool:
    if not cells:
        return False
    joined = " ".join(cells).lower()
    return (
        "dân" in joined
        or "gián" in joined
        or "civilian" in joined
        or "spy" in joined
        or joined.startswith("stt")
    )


def parse_csv_text(text: str) -> tuple[list[dict], int]:
    text = text.lstrip("\ufeff")
    if not text:
        return [], 0

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], 0

    start = 1 if looks_like_header(rows[0]) else 0
    pairs: list[dict] = []
    invalid = 0

    for row in rows[start:]:
        if not row:
            continue
        if len(row) >= 3:
            civilian, spy = row[1], row[2]
        elif len(row) == 2:
            civilian, spy = row[0], row[1]
        else:
            invalid += 1
            continue
        if not civilian or not spy:
            invalid += 1
            continue
        pairs.append({"civilian_word": civilian, "spy_word": spy})

    return pairs, invalid


def parse_comma_lines(text: str) -> tuple[list[dict], int]:
    pairs: list[dict] = []
    invalid = 0
    for line in text.splitlines():
        if not line:
            continue
        comma = line.find(",")
        if comma <= 0 or comma >= len(line) - 1:
            invalid += 1
            continue
        civilian = line[:comma]
        spy = line[comma + 1 :]
        if not civilian or not spy:
            invalid += 1
            continue
        pairs.append({"civilian_word": civilian, "spy_word": spy})
    return pairs, invalid


def parse_xlsx_bytes(data: bytes) -> tuple[list[dict], int]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if v is None else str(v) for v in row]
        if any(cells):
            rows.append(cells)
    wb.close()

    if not rows:
        return [], 0

    start = 1 if looks_like_header(rows[0]) else 0
    pairs: list[dict] = []
    invalid = 0

    for row in rows[start:]:
        if len(row) >= 3:
            civilian, spy = row[1], row[2]
        elif len(row) == 2:
            civilian, spy = row[0], row[1]
        else:
            invalid += 1
            continue
        if not civilian or not spy:
            invalid += 1
            continue
        pairs.append({"civilian_word": civilian, "spy_word": spy})

    return pairs, invalid


def parse_upload(filename: str, data: bytes) -> tuple[list[dict], int]:
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return parse_xlsx_bytes(data)
    text = data.decode("utf-8-sig")
    return parse_csv_text(text)